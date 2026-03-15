"""
Pre-process real data into county scores and bake into the GeoJSON.
Factors:
  1. Electricity price (state-level industrial rate, EIA) — lower is better
  2. Land cost (county-level median home value, Census ACS) — lower is better

Output: counties_scored.geojson with a 'score' property per feature.
"""
import json
import os
import openpyxl

# --- 1. Load electricity prices (state-level, industrial) ---
wb = openpyxl.load_workbook("eia_avgprice.xlsx")
ws = wb.active

# Find most recent year
years = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2] == "Total Electric Industry":
        years.add(row[0])
latest_year = max(years)

# state abbreviation -> industrial price (cents/kWh)
state_power_price = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] == latest_year and row[2] == "Total Electric Industry" and row[1] not in ("US", "DC"):
        state_abbr = row[1]
        price = row[5]  # Industrial column
        if isinstance(price, (int, float)):
            state_power_price[state_abbr] = price

# DC special case
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] == latest_year and row[2] == "Total Electric Industry" and row[1] == "DC":
        if isinstance(row[5], (int, float)):
            state_power_price["DC"] = row[5]

print(f"Loaded electricity prices for {len(state_power_price)} states (year {latest_year})")

# State FIPS -> state abbreviation mapping
FIPS_TO_ABBR = {
    "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA",
    "08": "CO", "09": "CT", "10": "DE", "11": "DC", "12": "FL",
    "13": "GA", "15": "HI", "16": "ID", "17": "IL", "18": "IN",
    "19": "IA", "20": "KS", "21": "KY", "22": "LA", "23": "ME",
    "24": "MD", "25": "MA", "26": "MI", "27": "MN", "28": "MS",
    "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI",
    "45": "SC", "46": "SD", "47": "TN", "48": "TX", "49": "UT",
    "50": "VT", "51": "VA", "53": "WA", "54": "WV", "55": "WI",
    "56": "WY", "72": "PR",
}

# --- 2. Load land costs (county-level median home value) ---
with open("census_home_values.json") as f:
    census_data = json.load(f)

# FIPS (state+county) -> median home value
county_home_value = {}
for row in census_data[1:]:  # skip header
    name, value, state_fips, county_fips = row
    fips = state_fips + county_fips
    if value and value != "null":
        county_home_value[fips] = int(value)

print(f"Loaded home values for {len(county_home_value)} counties")

# --- 3. Normalize to 0-100 scores (lower cost = higher score) ---
power_prices = list(state_power_price.values())
power_min, power_max = min(power_prices), max(power_prices)

home_values = [v for v in county_home_value.values() if v > 0]
home_min, home_max = min(home_values), max(home_values)
# Use 95th percentile as max to avoid outliers squishing everything
home_values_sorted = sorted(home_values)
home_p95 = home_values_sorted[int(len(home_values_sorted) * 0.95)]

print(f"Power price range: {power_min} - {power_max} cents/kWh")
print(f"Home value range: ${home_min:,} - ${home_max:,} (p95: ${home_p95:,})")


def normalize_inverse(value, vmin, vmax):
    """Lower value = higher score (0-100)."""
    if vmax == vmin:
        return 50
    clamped = max(vmin, min(vmax, value))
    return 100 * (1 - (clamped - vmin) / (vmax - vmin))


# --- 4. Load GeoJSON and compute scores ---
with open("counties.geojson") as f:
    geojson = json.load(f)

WEIGHT_POWER = 0.5
WEIGHT_LAND = 0.5

scored = 0
missing_power = 0
missing_land = 0

for feature in geojson["features"]:
    props = feature["properties"]
    state_fips = props["STATE"]
    county_fips = props["COUNTY"]
    fips = state_fips + county_fips
    state_abbr = FIPS_TO_ABBR.get(state_fips)

    # Power score
    power_score = None
    if state_abbr and state_abbr in state_power_price:
        power_score = normalize_inverse(state_power_price[state_abbr], power_min, power_max)
    else:
        missing_power += 1

    # Land score
    land_score = None
    if fips in county_home_value and county_home_value[fips] > 0:
        land_score = normalize_inverse(county_home_value[fips], home_min, home_p95)
    else:
        missing_land += 1

    # Combined score
    if power_score is not None and land_score is not None:
        score = WEIGHT_POWER * power_score + WEIGHT_LAND * land_score
    elif power_score is not None:
        score = power_score
    elif land_score is not None:
        score = land_score
    else:
        score = 50  # neutral fallback

    props["score"] = round(score, 1)
    props["power_price"] = state_power_price.get(state_abbr) if state_abbr else None
    props["home_value"] = county_home_value.get(fips)
    scored += 1

print(f"\nScored {scored} counties")
print(f"Missing power data: {missing_power}")
print(f"Missing land data: {missing_land}")

with open("counties_scored.geojson", "w") as f:
    json.dump(geojson, f)

print(f"Written to counties_scored.geojson ({os.path.getsize('counties_scored.geojson') / 1e6:.1f} MB)")
